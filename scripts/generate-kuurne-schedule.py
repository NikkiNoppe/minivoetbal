#!/usr/bin/env python3
"""Genereer Kuurne-speelschema in de 372 M-momenten.

Site-opzet: 11+11 ploegen × 3 rondes = 330 competitie + 21 beker (22 ploegen).
Excel: 372 speelmomenten (M). Rest blijft vrij.

Competitie: ronde 1 → 2 → 3, max. 1 speeldag overlap.
Beker: 21 wedstrijden op de 5 voorkeursweken (maandag eerst, geen vrijdag).
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path

EXCEL_DEFAULT = Path("/Users/nikki/Desktop/Voorstel dagen kalender.xlsx")
ROUNDS = 3
MAX_MATCHDAY_OVERLAP = 1  # huidige speeldag + 1 volgende
TEAMS_PER_POOL = 11
CUP_ROUNDS: list[tuple[str, int]] = [
    ("Voorronde", 6),
    ("Achtste finale", 8),
    ("Kwartfinale", 4),
    ("Halve finale", 2),
    ("Finale", 1),
]
PREFERRED_CUP_MONDAYS = [
    date(2026, 12, 14),
    date(2027, 4, 26),
    date(2027, 5, 24),
    date(2027, 6, 7),
    date(2027, 6, 21),
]
CUP_DAY_RANK = {"Maandag": 0, "Dinsdag": 1, "Donderdag": 2, "Woensdag": 3}


@dataclass(frozen=True)
class Slot:
    row: int
    day: date
    weekday: str
    start: str

    @property
    def iso_monday(self) -> date:
        return self.day - timedelta(days=self.day.weekday())


@dataclass
class Pairing:
    matchday: int
    home: str
    away: str
    slot: Slot | None = None
    kind: str = "competition"
    cup_round: str = ""


def round_of(matchday: int, matchdays_per_round: int) -> int:
    return (matchday - 1) // matchdays_per_round + 1


def time_str(tm: object) -> str:
    if isinstance(tm, datetime):
        return tm.strftime("%H:%M")
    if isinstance(tm, time):
        return tm.strftime("%H:%M")
    if isinstance(tm, timedelta):
        total = int(tm.total_seconds())
        return f"{total // 3600:02d}:{(total % 3600) // 60:02d}"
    return str(tm)[:5] if tm else ""


def load_workbook(path: Path):
    import openpyxl

    return openpyxl.load_workbook(path, data_only=True)


def load_m_slots(ws) -> list[Slot]:
    slots: list[Slot] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 8:
            continue
        code = row[0]
        if not code or str(code).strip().upper() != "M":
            continue
        dt, dayname, tm = row[2], row[3], row[4]
        if not isinstance(dt, datetime):
            continue
        start = time_str(tm)
        if not start:
            continue
        slots.append(Slot(row=i, day=dt.date(), weekday=str(dayname or ""), start=start))
    slots.sort(key=lambda s: (s.day, s.start, s.row))
    return slots


def load_pairings(ws) -> list[Pairing]:
    pairings: list[Pairing] = []
    for row in ws.iter_rows(values_only=True):
        if row[0] is None or not row[1] or not row[3]:
            continue
        pairings.append(
            Pairing(matchday=int(row[0]), home=str(row[1]).strip(), away=str(row[3]).strip())
        )
    pairings.sort(key=lambda p: (p.matchday, p.home, p.away))
    return pairings


def infer_matchdays_per_round(pairings: list[Pairing], rounds: int = ROUNDS) -> int:
    max_md = max(p.matchday for p in pairings)
    if max_md % rounds == 0:
        return max_md // rounds
    teams = {p.home for p in pairings} | {p.away for p in pairings}
    # Eén reeks-label (RK1/RK2); neem grootste pool.
    pools: dict[str, set[str]] = defaultdict(set)
    for t in teams:
        pools[t.split()[0]].add(t)
    n = max(len(v) for v in pools.values())
    return n if n % 2 else n - 1


def eligible_matchdays(
    incomplete: list[int],
    matchdays_per_round: int,
    overlap: int = MAX_MATCHDAY_OVERLAP,
) -> list[int]:
    """Vroegste open speeldag + max. `overlap` volgende (ook over rondegrens)."""
    if not incomplete:
        return []
    allowed = incomplete[: 1 + overlap]
    # Nooit meer dan 1 speeldag van de volgende ronde openen.
    first = incomplete[0]
    first_round = round_of(first, matchdays_per_round)
    out: list[int] = []
    extra_next_round = 0
    for md in allowed:
        r = round_of(md, matchdays_per_round)
        if r > first_round:
            extra_next_round += 1
            if extra_next_round > overlap:
                break
        out.append(md)
    return out


def team_free(
    home: str,
    away: str,
    slot: Slot,
    used_date: dict[date, set[str]],
    used_week: dict[date, dict[str, list[date]]],
    max_per_week: int = 2,
    min_gap_days: int = 2,
) -> bool:
    busy_d = used_date[slot.day]
    if home in busy_d or away in busy_d:
        return False
    week_map = used_week[slot.iso_monday]
    for team in (home, away):
        prevs = week_map.get(team, [])
        if len(prevs) >= max_per_week:
            return False
        if any(abs((slot.day - d).days) < min_gap_days for d in prevs):
            return False
    return True


def assign_schedule(
    slots: list[Slot],
    pairings: list[Pairing],
    matchdays_per_round: int,
    overlap: int = MAX_MATCHDAY_OVERLAP,
) -> list[Pairing]:
    by_md: dict[int, list[Pairing]] = defaultdict(list)
    for p in pairings:
        by_md[p.matchday].append(p)

    open_by_md: dict[int, list[Pairing]] = {md: list(ms) for md, ms in by_md.items()}
    used_date: dict[date, set[str]] = defaultdict(set)
    used_week: dict[date, dict[str, list[date]]] = defaultdict(dict)

    def incomplete_mds() -> list[int]:
        return sorted(md for md, ms in open_by_md.items() if ms)

    # Greedy: elk moment krijgt de vroegst mogelijke speeldag (met 1 overlap).
    for slot in slots:
        candidates = eligible_matchdays(incomplete_mds(), matchdays_per_round, overlap)
        placed = False
        for md in candidates:
            remaining = open_by_md[md]
            pick_at = next(
                (
                    i
                    for i, p in enumerate(remaining)
                    if team_free(p.home, p.away, slot, used_date, used_week)
                ),
                None,
            )
            if pick_at is None:
                continue
            p = remaining.pop(pick_at)
            p.slot = slot
            used_date[slot.day].update((p.home, p.away))
            week_map = used_week[slot.iso_monday]
            for team in (p.home, p.away):
                week_map.setdefault(team, []).append(slot.day)
            placed = True
            break
        if not placed:
            continue

    unplaced = [p for md in open_by_md for p in open_by_md[md]]
    if unplaced:
        raise RuntimeError(
            f"{len(unplaced)} wedstrijden niet geplaatst (eerste: speeldag "
            f"{unplaced[0].matchday} {unplaced[0].home}–{unplaced[0].away})."
        )
    return pairings


def summarize(pairings: list[Pairing], matchdays_per_round: int) -> str:
    lines: list[str] = []
    by_round: dict[int, list[Pairing]] = defaultdict(list)
    for p in pairings:
        by_round[round_of(p.matchday, matchdays_per_round)].append(p)

    for r in sorted(by_round):
        ps = [p for p in by_round[r] if p.slot]
        days = sorted({p.slot.day for p in ps})
        mds = sorted({p.matchday for p in ps})
        lines.append(
            f"Ronde {r}: speeldagen {mds[0]}–{mds[-1]}, "
            f"{days[0].isoformat()} t.e.m. {days[-1].isoformat()}, "
            f"{len(ps)} wedstrijden"
        )
        if r > 1:
            prev = [p for p in by_round[r - 1] if p.slot]
            prev_last = max(p.slot.day for p in prev)
            this_first = min(p.slot.day for p in ps)
            overlap_days = sorted(
                {p.slot.day for p in prev} & {p.slot.day for p in ps}
            )
            lines.append(
                f"  overlap met ronde {r - 1}: "
                f"{'geen gemeenschappelijke datum' if not overlap_days else ', '.join(d.isoformat() for d in overlap_days)}"
                f" (start ronde {r}: {this_first}, einde ronde {r - 1}: {prev_last})"
            )
    return "\n".join(lines)


def write_output(
    source: Path,
    dest: Path,
    pairings: list[Pairing],
    matchdays_per_round: int,
) -> None:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.load_workbook(source)
    cal = wb["Voorstel dagen kalender"]
    by_row = {p.slot.row: p for p in pairings if p.slot}

    cal.cell(1, 6, "Thuis")
    cal.cell(1, 7, "")
    cal.cell(1, 8, "Uit")
    cal.cell(1, 9, "Speeldag")
    cal.cell(1, 10, "Ronde")
    for col in range(6, 11):
        cal.cell(1, col).font = Font(bold=True)

    for row_idx, p in by_row.items():
        cal.cell(row_idx, 6, p.home)
        cal.cell(row_idx, 7, "–")
        cal.cell(row_idx, 8, p.away)
        if p.kind == "cup":
            cal.cell(row_idx, 9, p.cup_round)
            cal.cell(row_idx, 10, "Beker")
        else:
            cal.cell(row_idx, 9, p.matchday)
            cal.cell(row_idx, 10, round_of(p.matchday, matchdays_per_round))

    if "Gegenereerd schema" in wb.sheetnames:
        del wb["Gegenereerd schema"]
    out = wb.create_sheet("Gegenereerd schema")
    headers = ["Type", "Ronde", "Speeldag", "Datum", "Dag", "Uur", "Thuis", "Uit"]
    for i, h in enumerate(headers, 1):
        out.cell(1, i, h).font = Font(bold=True)
    ordered = sorted(
        (p for p in pairings if p.slot),
        key=lambda p: (p.slot.day, p.slot.start, p.matchday),
    )
    for i, p in enumerate(ordered, 2):
        s = p.slot
        assert s is not None
        if p.kind == "cup":
            out.cell(i, 1, "Beker")
            out.cell(i, 2, p.cup_round)
            out.cell(i, 3, "")
        else:
            out.cell(i, 1, "Competitie")
            out.cell(i, 2, round_of(p.matchday, matchdays_per_round))
            out.cell(i, 3, p.matchday)
        out.cell(i, 4, s.day.isoformat())
        out.cell(i, 5, s.weekday)
        out.cell(i, 6, s.start)
        out.cell(i, 7, p.home)
        out.cell(i, 8, p.away)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def circle_round_robin(teams: list[str], rounds: int = 3) -> list[Pairing]:
    """Circle-method; oneven ploegen krijgen een bye (n speeldagen/ronde)."""
    n = len(teams)
    has_bye = n % 2 == 1
    rotating: list[str | None] = (teams[1:] + [None]) if has_bye else list(teams[1:])
    days = n if has_bye else n - 1
    pairings: list[Pairing] = []
    md = 0
    for r in range(rounds):
        rot = list(rotating)
        for _ in range(days):
            md += 1
            ordered: list[str | None] = [teams[0], *rot]
            half = len(ordered) // 2
            for i in range(half):
                home, away = ordered[i], ordered[-1 - i]
                if home is None or away is None:
                    continue
                if r % 2 == 1:
                    home, away = away, home
                pairings.append(Pairing(md, home, away))
            rot = rot[-1:] + rot[:-1]
    return pairings


def default_competition_pairings() -> list[Pairing]:
    rk1 = circle_round_robin([f"RK1 {i}" for i in range(1, TEAMS_PER_POOL + 1)], ROUNDS)
    rk2 = circle_round_robin([f"RK2 {i}" for i in range(1, TEAMS_PER_POOL + 1)], ROUNDS)
    return rk1 + rk2


def reserve_cup_slots(
    slots: list[Slot],
    mondays: list[date] = PREFERRED_CUP_MONDAYS,
) -> tuple[list[Pairing], list[Slot]]:
    """21 bekerwedstrijden op voorkeursweken; vrijdag blijft competitie."""
    remaining = list(slots)
    cup: list[Pairing] = []
    if len(mondays) < len(CUP_ROUNDS):
        raise RuntimeError("Te weinig bekerweken voor 21 wedstrijden")

    for week_i, (label, need) in enumerate(CUP_ROUNDS):
        monday = mondays[week_i]
        week_slots = [
            s
            for s in remaining
            if s.iso_monday == monday and s.weekday in CUP_DAY_RANK
        ]
        week_slots.sort(key=lambda s: (CUP_DAY_RANK.get(s.weekday, 9), s.start, s.row))
        if len(week_slots) < need:
            raise RuntimeError(
                f"Beker {label}: week {monday} heeft {len(week_slots)} momenten, {need} nodig"
            )
        taken = week_slots[:need]
        taken_rows = {s.row for s in taken}
        remaining = [s for s in remaining if s.row not in taken_rows]
        for j, slot in enumerate(taken, 1):
            cup.append(
                Pairing(
                    matchday=0,
                    home="BEKER",
                    away=f"{label} {j}",
                    slot=slot,
                    kind="cup",
                    cup_round=label,
                )
            )
    return cup, remaining


def self_test() -> None:
    """11+11 × 3 rondes = 330; oneven reeks → 11 speeldagen/ronde + bye."""
    start = date(2026, 8, 24)
    slots: list[Slot] = []
    row = 8
    for w in range(50):
        monday = start + timedelta(weeks=w)
        thursday = monday + timedelta(days=3)
        for d, name in ((monday, "Maandag"), (thursday, "Donderdag")):
            for hour in (18, 19, 20, 21):
                slots.append(Slot(row=row, day=d, weekday=name, start=f"{hour:02d}:00"))
                row += 1

    pairings = default_competition_pairings()
    assert len(pairings) == 330
    placed = assign_schedule(slots, pairings, matchdays_per_round=11, overlap=1)
    assert all(p.slot for p in placed)

    by_date: dict[date, set[int]] = defaultdict(set)
    for p in placed:
        assert p.slot
        by_date[p.slot.day].add(p.matchday)
    for d, mds in by_date.items():
        assert max(mds) - min(mds) <= 1, f"{d}: {sorted(mds)}"

    r1_core = [p for p in placed if p.matchday <= 10]
    r2 = [p for p in placed if 12 <= p.matchday <= 22]
    last_r1_core = max(p.slot.day for p in r1_core if p.slot)
    first_r2 = min(p.slot.day for p in r2 if p.slot)
    assert first_r2 >= last_r1_core, (first_r2, last_r1_core)
    print("self-test ok")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("excel", nargs="?", type=Path, default=EXCEL_DEFAULT)
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output-xlsx (default: <excel>-schema.xlsx)",
    )
    parser.add_argument(
        "--from-sheet1",
        action="store_true",
        help="Gebruik Sheet1-pairings i.p.v. 11+11 × 3 rondes",
    )
    parser.add_argument("--self-test", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        self_test()
        return

    excel = args.excel
    if not excel.exists():
        print(f"Excel niet gevonden: {excel}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(excel)
    slots = load_m_slots(wb["Voorstel dagen kalender"])
    print(f"{len(slots)} M-speelmomenten (Excel)")

    cup, rest = reserve_cup_slots(slots)
    print(f"Beker: {len(cup)} wedstrijden op {len(PREFERRED_CUP_MONDAYS)} weken")

    if args.from_sheet1:
        pairings = load_pairings(wb["Sheet1"])
        per_round = infer_matchdays_per_round(pairings)
    else:
        pairings = default_competition_pairings()
        per_round = TEAMS_PER_POOL if TEAMS_PER_POOL % 2 else TEAMS_PER_POOL - 1
    print(f"Competitie: {len(pairings)} wedstrijden, {per_round} speeldagen/ronde")

    placed = assign_schedule(rest, pairings, per_round, overlap=MAX_MATCHDAY_OVERLAP)
    print(summarize(placed, per_round))

    dest = args.output or excel.with_name(f"{excel.stem}-schema{excel.suffix}")
    write_output(excel, dest, cup + placed, per_round)
    used = sum(1 for p in cup + placed if p.slot)
    print(
        f"Schreef {dest} ({len(placed)} competitie + {len(cup)} beker = {used} / {len(slots)} momenten, "
        f"{len(slots) - used} vrij)"
    )


if __name__ == "__main__":
    main()
